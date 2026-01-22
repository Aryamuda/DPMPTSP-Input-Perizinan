import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
import os
from datetime import datetime

# Page config
st.set_page_config(page_title="Data Processor", layout="wide", page_icon="📊")

# Initialize session state
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'stacked_data' not in st.session_state:
    st.session_state.stacked_data = None
if 'column_mappings' not in st.session_state:
    st.session_state.column_mappings = {}

# Utility functions
def extract_patterns(text, column_name=''):
    """Extract name, company, vessel, email, phone, NIK, NPWP, NIB from text"""
    if pd.isna(text) or not isinstance(text, str):
        return {}
    
    result = {}
    
    # Clean text for processing
    text_clean = text.strip()
    col_upper = column_name.upper()
    
    # Person/Company Name (usually at the beginning, before any identifiers)
    
    # Pattern 1: Name/KM: or Name/KM or Name KM: format
    name_slash_km = re.search(r'^([A-Z][A-Z\s\.\,\&\-\']+?)\s*(?:/|(?=KM:))\s*(?=KM[:\s\.])', text_clean, re.IGNORECASE)
    if name_slash_km:
        name = name_slash_km.group(1).strip()
        # Check if company (allow multiple spaces)
        if re.match(r'^(PT|CV|UD|PD)\s*[\.\s]', name, re.IGNORECASE):
            result['company_name'] = name
        else:
            result['person_name'] = name
    # Pattern 2: Name before KM
    elif not result.get('person_name') and not result.get('company_name'):
        name_before_km = re.search(r'^([A-Z][A-Z\s\.\,\&\-\']+?)\s+(?=KM[:\s\.])', text_clean, re.IGNORECASE)
        if name_before_km:
            name = name_before_km.group(1).strip()
            if re.match(r'^(PT|CV|UD|PD)\s*[\.\s]', name, re.IGNORECASE):
                result['company_name'] = name
            else:
                result['person_name'] = name
    
    # Pattern 3: Name/Something without KM
    if not result.get('person_name') and not result.get('company_name'):
        name_slash_other = re.search(r'^([A-Z][A-Z\s\.]+?)/([A-Z\s\d]+?)(?=\s+|NPWP|NIK|EMAIL|TELP|$)', text_clean, re.IGNORECASE)
        if name_slash_other:
            result['person_name'] = name_slash_other.group(1).strip()
            result['vessel_name'] = name_slash_other.group(2).strip()
    
    # Pattern 4: Regular name pattern (simple format like Juli)
    if not result.get('person_name') and not result.get('company_name'):
        name_pattern = r'^([A-Z][A-Z\s\.\,\&\-\'\d]+?)(?=\s{2,}|NIK|NPWP|NIB|EMAIL|TELP|TLP|NOMOR|STATUS|$)'
        name_match = re.search(name_pattern, text_clean, re.IGNORECASE | re.MULTILINE)
        if name_match:
            name = name_match.group(1).strip()
            if re.match(r'^(PT|CV|UD|PD)\s*[\.\s]', name, re.IGNORECASE):
                result['company_name'] = name
            else:
                result['person_name'] = name
    
    # Vessel/Boat Name (KM: or KM. or KM space)
    vessel_pattern = r'(?:/\s*)?KM[:\s\.]([A-Z\s\d]+?)(?=\s{2,}|NIK|NPWP|NIB|EMAIL|TELP|TLP|STATUS|$)'
    vessel_match = re.search(vessel_pattern, text_clean, re.IGNORECASE)
    if vessel_match:
        result['vessel_name'] = f"KM {vessel_match.group(1).strip()}"
    
    # KBLI pattern (code + description, e.g., "50133,  PENGANGKUTAN IKAN")
    kbli_pattern = re.search(r'(\d{5})\s*,\s*([A-Z\s]+)', text_clean, re.IGNORECASE)
    if kbli_pattern:
        result['kbli'] = f"{kbli_pattern.group(1)} - {kbli_pattern.group(2).strip()}"
    
    # Email pattern (handle EMAIL: EMAIL; or stuck together)
    email_with_delimiter = re.search(r'EMAIL[;:\s]*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', text_clean, re.IGNORECASE)
    if email_with_delimiter:
        result['email'] = email_with_delimiter.group(1)
    else:
        # Try without EMAIL keyword
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails = re.findall(email_pattern, text)
        if emails:
            result['email'] = emails[0]
    
    # === SMART NUMBER EXTRACTION ===
    
    # 1. Try labeled extraction first (NPWP.xxx, NIK.xxx, NIB.xxx, TELP.xxx)
    npwp_labeled = re.search(r'NPWP[:\.\s]+(\d{15})', text_clean, re.IGNORECASE)
    if npwp_labeled:
        result['npwp'] = npwp_labeled.group(1)
    
    nik_labeled = re.search(r'NIK[:\.\s]+(\d{16})', text_clean, re.IGNORECASE)
    if nik_labeled:
        result['nik'] = nik_labeled.group(1)
    
    nib_labeled = re.search(r'NIB[:\.\s]+(\d{13})', text_clean, re.IGNORECASE)
    if nib_labeled:
        result['nib'] = nib_labeled.group(1)
    
    # Phone with label (handle stuck together, quotes, TELPON/NOMOR variants)
    phone_labeled = re.search(r'(?:TELPON|TELP|TLP|HP|PHONE|NOMOR)[:\.\'\s]*(\+?62|0)?[\s-]?(\d{8,13})', text_clean, re.IGNORECASE)
    if phone_labeled:
        # Clean up quotes and extra characters
        phone_parts = [p for p in phone_labeled.groups() if p]
        result['phone'] = ''.join(phone_parts).replace("'", "").replace('"', '')
    
    # 2. Handle multi-separator values (/, |, comma, space, newline)
    # Split by various separators and try to match by column hints
    separators = [r'/', r'\|', r',', r'\s{2,}', r'\n', r';']
    
    for separator in separators:
        if re.search(separator, text_clean) and re.search(separator.replace('\\', ''), col_upper):
            parts = re.split(separator, text_clean)
            col_parts = re.split(separator.replace('\\', ''), col_upper)
            
            # Ensure we process all parts even if lengths don't match
            max_len = max(len(parts), len(col_parts))
            
            for i in range(max_len):
                part = parts[i].strip() if i < len(parts) else ''
                col_hint = col_parts[i].strip() if i < len(col_parts) else ''
                
                if not part:
                    continue
                
                # Extract all digit sequences from part
                numbers = re.findall(r'\d+', part)
                
                for num in numbers:
                    # Match by column hint and number length (with some flexibility)
                    if 'NPWP' in col_hint and 'npwp' not in result:
                        # NPWP should be 15 digits, but be flexible (14-16)
                        if 14 <= len(num) <= 16:
                            result['npwp'] = num
                    elif 'NIK' in col_hint and 'nik' not in result:
                        # NIK should be 16 digits, but be flexible (15-17)
                        if 15 <= len(num) <= 17:
                            result['nik'] = num
                    elif 'NIB' in col_hint and 'nib' not in result:
                        # NIB should be 13 digits, but be flexible (12-14)
                        if 12 <= len(num) <= 14:
                            result['nib'] = num
                    elif any(x in col_hint for x in ['TELP', 'PHONE', 'HP', 'NOMOR']) and 'phone' not in result:
                        if 10 <= len(num) <= 15:
                            result['phone'] = num
            
            # If we found matches, don't try other separators
            if any(k in result for k in ['npwp', 'nik', 'nib', 'phone']):
                break
    
    # 3. Fallback: Extract all numbers and classify by length (if not already found)
    # This ensures we get ALL matching numbers, not just the first one
    all_numbers = re.findall(r'\b\d{10,16}\b', text_clean)
    
    for num in all_numbers:
        length = len(num)
        
        # NPWP = 15 digits (rare, high priority)
        if length == 15 and 'npwp' not in result:
            result['npwp'] = num
        
        # NIK = 16 digits  
        elif length == 16 and 'nik' not in result:
            result['nik'] = num
        
        # NIB = 13 digits (differentiate from phone)
        elif length == 13 and 'nib' not in result:
            # If column name has NIB, definitely NIB
            if 'NIB' in col_upper:
                result['nib'] = num
            # If doesn't start with 0/62 and no phone indicators, likely NIB
            elif not num.startswith(('0', '62')) and not any(x in col_upper for x in ['TELP', 'PHONE', 'HP']):
                result['nib'] = num
            # Otherwise might be phone
            elif 'phone' not in result:
                result['phone'] = num
        
        # Phone = 10-13 digits, usually starts with 0 or 62
        elif 10 <= length <= 13 and 'phone' not in result:
            if num.startswith(('0', '62', '8')) or any(x in col_upper for x in ['TELP', 'PHONE', 'HP', 'NOMOR']):
                result['phone'] = num
    
    return result

def get_excel_column_letter(idx):
    """Convert column index to Excel letter (0 -> A, 1 -> B, etc.)"""
    result = ""
    while idx >= 0:
        result = chr(65 + (idx % 26)) + result
        idx = idx // 26 - 1
    return result

def highlight_cells(df, start_row, start_col):
    """Create Excel with highlighted starting position"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Highlight the starting cell
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell = worksheet.cell(row=start_row+1, column=start_col+1)
        cell.fill = fill
        
    return output.getvalue()

# Main app
st.title("Data Processor Untuk Excel/CSV")
st.markdown("---")

# Step 1: File Upload
st.header("1️⃣ Upload File")
uploaded_file = st.file_uploader("Upload Excel or CSV file", type=['xlsx', 'xls', 'csv'])

if uploaded_file:
    # Read file
    if uploaded_file.name.endswith('.csv'):
        df_raw = pd.read_csv(uploaded_file, header=None)
        sheet_names = ['Sheet1']
        selected_sheet = 'Sheet1'
    else:
        xl_file = pd.ExcelFile(uploaded_file)
        sheet_names = xl_file.sheet_names
        
        # Step 2: Sheet Selection
        st.header("2️⃣ Pilih Sheet")
        selected_sheet = st.selectbox("Pilih sheet untuk diproses:", sheet_names)
        df_raw = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)
    
    st.success(f"Loaded: **{uploaded_file.name}** | Sheet: **{selected_sheet}**")
    
    # Step 3: Define Starting Position
    st.header("3️⃣ Atur Posisi Awal")
    st.info("Pilih di mana tabel data sebenarnya dimulai (baris header dan kolom pertama)")
    
    col1, col2 = st.columns(2)
    with col1:
        start_row = st.number_input("Baris Awal (0-indexed)", min_value=0, max_value=len(df_raw)-1, value=5)
    with col2:
        start_col = st.number_input("Kolom Awal (0-indexed)", min_value=0, max_value=len(df_raw.columns)-1, value=0)
    
    # Preview with highlight
    st.subheader("📋 Preview (20 Baris Awal)")
    preview_df = df_raw.head(20).copy()
    
    # Style the dataframe to highlight selected cell
    def highlight_start(row):
        styles = [''] * len(row)
        return styles
    
    # Create visual preview
    st.dataframe(preview_df, width='stretch', height=400)
    st.caption(f"🟡 Posisi Awal: Baris {start_row}, Kolom {start_col} (Excel: {get_excel_column_letter(start_col)}{start_row+1})")
    
    # Extract data from starting position
    df_working = df_raw.iloc[start_row:, start_col:].reset_index(drop=True)
    
    # Check for multi-level headers
    st.subheader("🔍 Header Structure")
    has_nested_headers = st.checkbox("Some columns have nested/multi-level headers (e.g., 'KBLI DAN JENIS IZIN' splits into 'KBLI' and 'JENIS IZIN')", value=False)
    
    if has_nested_headers:
        st.info("📋 Multi-level header detected. We'll combine parent names with child names where applicable.")
        
        parent_header_row = st.number_input("Parent header row (relative to start):", 
                                            min_value=0, max_value=min(5, len(df_working)-1), value=0)
        child_header_row = st.number_input("Child header row (relative to start):", 
                                           min_value=0, max_value=min(5, len(df_working)-1), value=1)
        
        # Get parent and child headers
        parent_headers = df_working.iloc[parent_header_row]
        child_headers = df_working.iloc[child_header_row]
        
        # Forward fill parent headers (to handle merged cells)
        parent_filled = parent_headers.fillna(method='ffill')
        
        # Combine headers intelligently
        combined_headers = []
        for i, (parent, child) in enumerate(zip(parent_filled, child_headers)):
            parent_str = str(parent).strip() if pd.notna(parent) and str(parent).strip() not in ['nan', ''] else ''
            child_str = str(child).strip() if pd.notna(child) and str(child).strip() not in ['nan', ''] else ''
            
            # If both exist and parent is different from child (meaning it's a true parent-child relationship)
            if parent_str and child_str:
                # Check if parent appears in multiple consecutive columns (indicating it's a spanning header)
                # Compare with previous column to see if it's a continuation
                if i > 0 and str(parent_filled.iloc[i-1]).strip() == parent_str:
                    # This is a child column under a parent header
                    combined_headers.append(f"{parent_str} - {child_str}")
                elif i < len(parent_filled)-1 and str(parent_filled.iloc[i+1]).strip() == parent_str:
                    # This is the first child column under a parent header
                    combined_headers.append(f"{parent_str} - {child_str}")
                else:
                    # Parent only appears once, so it's not really a spanning header
                    # Use child or parent (whichever is more meaningful)
                    combined_headers.append(child_str if child_str else parent_str)
            # If only child exists
            elif child_str:
                combined_headers.append(child_str)
            # If only parent exists
            elif parent_str:
                combined_headers.append(parent_str)
            # If neither, use index
            else:
                combined_headers.append(f"Column_{i}")
        
        df_working.columns = combined_headers
        df_working = df_working.iloc[child_header_row+1:].reset_index(drop=True)
        
        st.success(f"✅ Processed {len(combined_headers)} column headers")
        with st.expander("📋 Preview Combined Headers"):
            for i, col in enumerate(combined_headers[:15]):
                st.text(f"{i+1}. {col}")
            if len(combined_headers) > 15:
                st.text(f"... and {len(combined_headers)-15} more columns")
    else:
        # Use first row as header
        header_row = st.number_input("Which row (relative to start) contains column names?", 
                                      min_value=0, max_value=min(5, len(df_working)-1), value=0)
        
        df_working.columns = df_working.iloc[header_row]
        df_working = df_working.iloc[header_row+1:].reset_index(drop=True)
        
        # Clean up column names
        df_working.columns = [str(col).strip() if pd.notna(col) else f"Unnamed_{i}" 
                              for i, col in enumerate(df_working.columns)]
    
    st.success(f"✅ Extracted {len(df_working)} rows, {len(df_working.columns)} columns")
    
    # Step 4: Select Columns
    st.header("4️⃣ Pilih Kolom untuk Diekstrak")
    
    all_columns = list(df_working.columns)
    selected_columns = st.multiselect(
        "Pilih kolom yang ingin disimpan:",
        options=all_columns,
        default=all_columns[:min(5, len(all_columns))]
    )
    
    if selected_columns:
        df_selected = df_working[selected_columns].copy()
        
        # Step 5: Column Splitting & Extraction
        st.header("5️⃣ Extract Data from Columns")
        st.info("Select columns that contain multiple data types to extract")
        
        columns_to_split = st.multiselect(
            "Which columns contain mixed data to extract?",
            options=selected_columns,
            help="Choose columns that have multiple values mixed together"
        )
        
        # Process splitting
        if columns_to_split:
            for col in columns_to_split:
                st.subheader(f"🔍 Extracting from: **{col}**")
                
                # Show sample data
                sample_data = df_selected[col].dropna().head(3)
                with st.expander("👁️ View sample data"):
                    for idx, val in enumerate(sample_data):
                        st.text(f"  {idx+1}. {val}")
                
                # Let user choose what to extract
                st.markdown("**Select what to extract:**")
                
                extraction_options = {
                    'name': '👤 Name (person or company)',
                    'vessel_name': '🚢 Vessel/Boat name (KM.)',
                    'kbli': '📋 KBLI (code + description)',
                    'email': '📧 Email',
                    'phone': '📞 Phone number',
                    'nik': '🆔 NIK (16 digits)',
                    'npwp': '💼 NPWP (15 digits)',
                    'nib': '🏢 NIB (13 digits)',
                }
                
                selected_extractions = []
                cols_choice = st.columns(4)
                for idx, (key, label) in enumerate(extraction_options.items()):
                    with cols_choice[idx % 4]:
                        if st.checkbox(label, key=f"extract_{col}_{key}"):
                            selected_extractions.append(key)
                
                if selected_extractions:
                    # Extract patterns (pass column name for better detection)
                    extracted = df_selected[col].apply(lambda x: extract_patterns(x, col))
                    
                    # Create new columns for selected extraction types only
                    for key in selected_extractions:
                        if key == 'name':
                            # Special handling for name - combine person and company
                            person_name = extracted.apply(lambda x: x.get('person_name', None))
                            company_name = extracted.apply(lambda x: x.get('company_name', None))
                            
                            new_col_name = f"{col}_name"
                            df_selected[new_col_name] = person_name.where(person_name.notna(), company_name)
                            
                            non_null = df_selected[new_col_name].notna().sum()
                            if non_null > 0:
                                st.success(f"  ✓ Extracted {non_null} NAME values → **{new_col_name}**")
                            else:
                                st.warning(f"  ⚠️ No NAME found in this column")
                        else:
                            new_col_name = f"{col}_{key}"
                            df_selected[new_col_name] = extracted.apply(lambda x: x.get(key, None))
                            
                            # Show extraction stats
                            non_null = df_selected[new_col_name].notna().sum()
                            total = len(df_selected[new_col_name])
                            if non_null > 0:
                                display_name = key.upper().replace('_', ' ')
                                st.success(f"  ✓ Extracted {non_null}/{total} {display_name} values → **{new_col_name}**")
                                
                                # Show sample of extracted values
                                samples = df_selected[new_col_name].dropna().head(3).tolist()
                                if samples:
                                    st.caption(f"     Sample: {', '.join(map(str, samples[:2]))}...")
                            else:
                                st.warning(f"  ⚠️ No {key.upper()} found in this column - check data format")
                else:
                    st.warning("  ⚠️ No extraction types selected for this column")
        
        # Step 6: Map to Standard Columns
        st.header("6️⃣ Map ke Struktur Kolom Standar")
        st.info("Pemetaan kolom data Anda ke struktur output standar")
        
        # Define standard columns
        standard_columns = ['NIB', 'Nama', 'KBLI', 'Alamat', 'NPWP', 'Nomor', 'Email']
        
        # Get all available columns (including extracted ones)
        available_columns = ['[None]'] + list(df_selected.columns)
        
        # Create mapping interface
        column_mapping = {}
        
        cols = st.columns(2)
        for idx, std_col in enumerate(standard_columns):
            with cols[idx % 2]:
                mapped_col = st.selectbox(
                    f"**{std_col}** ← Map dari:",
                    options=available_columns,
                    key=f"map_{std_col}",
                    help=f"Pilih kolom yang akan dipetakan ke {std_col}"
                )
                if mapped_col != '[None]':
                    column_mapping[std_col] = mapped_col
        
        # Apply mapping and create standardized dataframe
        df_standardized = pd.DataFrame()
        
        for std_col in standard_columns:
            if std_col in column_mapping:
                source_col = column_mapping[std_col]
                df_standardized[std_col] = df_selected[source_col]
            else:
                df_standardized[std_col] = None
        
        # FIRST: Fill original null/empty values with "-"
        df_standardized = df_standardized.fillna("-")
        df_standardized = df_standardized.replace('', "-")
        df_standardized = df_standardized.replace('nan', "-")
        
        # SECOND: Sort by Nama to group same names together
        if 'Nama' in df_standardized.columns:
            # Sort: real names first (not "-"), then by name alphabetically
            df_standardized['_sort_key'] = df_standardized['Nama'].apply(lambda x: (x == '-', x))
            df_standardized = df_standardized.sort_values('_sort_key').drop('_sort_key', axis=1).reset_index(drop=True)
            st.info(f"ℹ️ Data sorted by Nama to group duplicates together")
        
        # THEN: Blank duplicate names (set to empty string, not "-")
        if 'Nama' in df_standardized.columns:
            # Find duplicates and blank them (keep first occurrence)
            mask = (df_standardized['Nama'] != '-') & (df_standardized['Nama'] != '')
            duplicate_mask = mask & df_standardized.duplicated(subset=['Nama'], keep='first')
            
            duplicate_count = duplicate_mask.sum()
            if duplicate_count > 0:
                df_standardized.loc[duplicate_mask, 'Nama'] = ''
                st.warning(f"⚠️ Blanked {duplicate_count} duplicate names (grouped by sorting first)")
        
        st.success(f"✅ Memetakan {len(column_mapping)} kolom ke struktur standar")
        
        # Show mapping summary
        with st.expander("📋 Ringkasan Pemetaan"):
            for std_col in standard_columns:
                if std_col in column_mapping:
                    st.text(f"✓ {std_col} ← {column_mapping[std_col]}")
                else:
                    st.text(f"✗ {std_col} ← [Not mapped]")
        
        # Use standardized dataframe
        df_selected = df_standardized
        
        # Add month column
        st.header("7️⃣ Tambah Informasi Bulan")
        col1, col2 = st.columns(2)
        with col1:
            month = st.text_input("Month", value=selected_sheet)
        with col2:
            year = st.number_input("Year", value=2025, min_value=2000, max_value=2100)
        
        df_selected.insert(0, 'BULAN', f"{month} {year}")
        
        # Preview processed data
        st.subheader("👀 Pratinjau Data yang Diproses")
        st.dataframe(df_selected, width='stretch', height=400)
        st.caption(f"Total baris bulan ini: {len(df_selected)}")
        
        # Store current month data in session state
        st.session_state.processed_data = df_selected.copy()
        
        # Stack with existing data
        st.header("8️⃣ Stack Data (Multi-Month)")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            if st.button("➕ Add to Stack", type="primary", width='stretch'):
                if st.session_state.stacked_data is None:
                    st.session_state.stacked_data = df_selected.copy()
                    st.success(f"✅ First month added! Total: {len(st.session_state.stacked_data)} rows")
                else:
                    st.session_state.stacked_data = pd.concat(
                        [st.session_state.stacked_data, df_selected], 
                        ignore_index=True
                    )
                    st.success(f"✅ Month added! Total: {len(st.session_state.stacked_data)} rows")
        
        with col2:
            if st.button("🗑️ Clear Stack", width='stretch'):
                st.session_state.stacked_data = None
                st.success("Stack cleared!")
                st.rerun()
        
        # Show stacked data preview
        if st.session_state.stacked_data is not None:
            st.subheader("📚 Stacked Data (All Months)")
            
            # Show summary by month
            month_counts = st.session_state.stacked_data['BULAN'].value_counts().sort_index()
            st.info(f"**Months in stack:** {', '.join(month_counts.index.tolist())}")
            
            # Show month breakdown
            cols = st.columns(len(month_counts))
            for idx, (month, count) in enumerate(month_counts.items()):
                with cols[idx]:
                    st.metric(month, f"{count} rows")
            
            # Show ALL data (not just first 20)
            st.dataframe(st.session_state.stacked_data, width='stretch', height=500)
            st.caption(f"Showing ALL {len(st.session_state.stacked_data)} rows from {len(month_counts)} month(s)")
        
        # Step 9: Save & Download
        st.header("9️⃣ Simpan & Download")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📥 Download Current Month")
            
            # Convert to Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_selected.to_excel(writer, index=False, sheet_name='Data')
            
            st.download_button(
                label="⬇️ Download Current Month",
                data=output.getvalue(),
                file_name=f"processed_{selected_sheet}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )
        
        with col2:
            st.subheader("📦 Download ALL Stacked Data")
            
            if st.session_state.stacked_data is not None:
                # Convert stacked data to Excel
                output_stacked = BytesIO()
                with pd.ExcelWriter(output_stacked, engine='openpyxl') as writer:
                    st.session_state.stacked_data.to_excel(writer, index=False, sheet_name='All Data')
                
                st.download_button(
                    label="⬇️ Download All Months",
                    data=output_stacked.getvalue(),
                    file_name=f"all_months_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch',
                    type="primary"
                )
            else:
                st.info("No stacked data yet. Process and add months first.")
else:
    st.info("👆 Upload a file to get started")
    
    # Show example
    with st.expander("ℹ️ How to use this app"):
        st.markdown("""
        ### Step-by-step guide:
        
        1. **Upload** your Excel or CSV file
        2. **Select** the sheet you want to process (if Excel has multiple sheets)
        3. **Define** where your actual data table starts (row and column)
        4. **Choose** which columns to extract
        5. **Extract** data from mixed columns (email, phone, NIK, NPWP will be auto-detected)
        6. **Rename** columns if needed
        7. **Add** month/year information
        8. **Download** processed data or **merge** with previous months
        
        ### Features:
        - ✅ Automatic extraction of email, phone, NIK, NPWP
        - ✅ Handle merged headers and complex table structures
        - ✅ Visual preview of data
        - ✅ Month tracking for multi-month merging
        - ✅ Column renaming and mapping
        """)

# Sidebar info
with st.sidebar:
    st.header("ℹ️ About")
    st.markdown("""
    **Data Processor v1.0**
    
    Built for processing maritime and fisheries sector reports.
    
    ---
    
    **Supported formats:**
    - Excel (.xlsx, .xls)
    - CSV (.csv)
    
    **Auto-detection:**
    - � Person names
    - 🏢 Company names (PT., CV., etc.)
    - 🚢 Vessel names (KM.)
    - �📧 Email addresses
    - 📞 Phone numbers (ID format)
    - 🆔 NIK (16 digits)
    - 💼 NPWP (15 digits)
    """)
    
    if st.session_state.processed_data is not None:
        st.success("✅ Data processed!")
        st.metric("Current Month Rows", len(st.session_state.processed_data))
        
        if st.session_state.stacked_data is not None:
            st.metric("Total Stacked Rows", len(st.session_state.stacked_data))
            months = st.session_state.stacked_data['BULAN'].nunique()
            st.metric("Months in Stack", months)
